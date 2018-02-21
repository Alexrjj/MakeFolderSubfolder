import os
import openpyxl
import time
import shutil

wb = openpyxl.load_workbook('pastas.xlsx')

for sheet in wb.worksheets:
    path = sheet['A2'].value
    for row in sheet.iter_cols(min_row=2, min_col=2, max_col=2):
        print('Change the date in xlsm file to ' + sheet['A2'].value)
        os.startfile('FluxoObraPadrao4.0.xlsm')
        for x in range(0, 1000):
            try:
                xls = open('FluxoObraPadrao4.0.xlsm', 'r+')
                if xls:
                    xls.close()
                    break
            except IOError:
                time.sleep(1)
                continue
        for cell in row:
            os.makedirs(os.path.join(path, str(cell.value)))
            shutil.copy2('FluxoObraPadrao4.0.xlsm', os.path.join(path, str(cell.value)))
            os.rename(os.path.join(path, str(cell.value), 'FluxoObraPadrao4.0.xlsm'), (os.path.join(path, str(cell.value), str(cell.value) + '.xlsm')))
            '''
            Código abaixo não funcional, pois ao inserir o número da SOB na célula específica, e salvar o arquivo,
            perde-se boa parte da formatação do mesmo.
            wb1 = openpyxl.load_workbook(os.path.join(path, str(cell.value), str(cell.value) + '.xlsm'), read_only=False, keep_vba=True)
            sheet2 = wb1['Ficha_Solicitação_Material']
            sheet2['H5'] = str(cell.value)
            wb1.save(str(cell.value + '.xlsm'))
            '''
            time.sleep(3)